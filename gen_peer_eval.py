from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()
wb.remove(wb.active)

TEAM = [
    ("Daniyal Ahmed",       "SP23-BCT-011"),
    ("Shaheer Khalid",      "SP23-BCT-048"),
    ("Maaz Malik",          "SP23-BCT-025"),
    ("Rana Mutahhar Ahmed", "SP23-BCT-045"),
]

SHEET_NAMES = ["Eval_Daniyal", "Eval_Shaheer", "Eval_Maaz", "Eval_Rana"]

NAVY_FILL  = PatternFill("solid", fgColor="1F3864")
GREY_FILL  = PatternFill("solid", fgColor="D9D9D9")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
BLUE_FILL  = PatternFill("solid", fgColor="DDEEFF")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
INFO_FILL  = PatternFill("solid", fgColor="E8EFF8")
INSTR_FILL = PatternFill("solid", fgColor="FFF2CC")
SIG_FILL   = PatternFill("solid", fgColor="F2F2F2")

thin = Side(style="thin", color="AAAAAA")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def apply_border(cell):
    cell.border = THIN_BORDER

def hdr_cell(cell, text):
    cell.value     = text
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell.fill      = NAVY_FILL
    cell.alignment = CENTER
    apply_border(cell)

for idx, (ev_name, ev_reg) in enumerate(TEAM):
    ws = wb.create_sheet(title=SHEET_NAMES[idx])

    col_widths = {"A": 26, "B": 16, "C": 18, "D": 16,
                  "E": 14, "F": 18, "G": 13, "H": 14, "I": 30}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 6
    ws.row_dimensions[4].height = 28
    ws.row_dimensions[5].height = 6
    ws.row_dimensions[6].height = 30
    for r in range(7, 11):
        ws.row_dimensions[r].height = 22

    # Row 1 - Title
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value     = "Peer Evaluation Form \u2014 SecureVault Team"
    c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=14)
    c.fill      = NAVY_FILL
    c.alignment = CENTER

    # Row 2 - Info bar
    ws.merge_cells("A2:I2")
    c = ws["A2"]
    c.value = (
        "Course: CYC386 \u2014 Secure Software Design and Development   |   "
        "Evaluator: " + ev_name + "   |   Reg No: " + ev_reg + "   |   "
        "Instructor: Engr. Muhammad Ahmad Nawaz   |   Date: 2026-04-08"
    )
    c.font      = Font(name="Arial", size=10, color="1F3864")
    c.fill      = INFO_FILL
    c.alignment = CENTER

    # Row 4 - Instructions
    ws.merge_cells("A4:I4")
    c = ws["A4"]
    c.value = (
        "Instructions: Rate each team member (excluding yourself) on a scale of 1-10 "
        "for each criterion. Be honest and fair."
    )
    c.font      = Font(name="Arial", italic=True, size=10, color="595959")
    c.fill      = INSTR_FILL
    c.alignment = LEFT

    # Row 6 - Headers
    headers = [
        "Team Member", "Reg No.", "Contribution\nto Code",
        "Documentation", "Participation", "Technical\nKnowledge",
        "Teamwork", "TOTAL\n(/50)", "Comments"
    ]
    for col_i, hdr in enumerate(headers, start=1):
        hdr_cell(ws.cell(row=6, column=col_i), hdr)

    # Data validation
    dv = DataValidation(
        type="whole", operator="between",
        formula1=1, formula2=10,
        showDropDown=False,
        error="Please enter an integer between 1 and 10.",
        errorTitle="Invalid Score",
        prompt="Enter a score from 1 to 10.",
        promptTitle="Score"
    )
    ws.add_data_validation(dv)

    for row_i, (m_name, m_reg) in enumerate(TEAM):
        row = 7 + row_i
        is_self = (m_name == ev_name)
        row_fill = GREY_FILL if is_self else (BLUE_FILL if row_i % 2 == 0 else WHITE_FILL)

        # Name
        c = ws.cell(row=row, column=1)
        c.value     = m_name
        c.font      = Font(name="Arial", italic=True, color="808080", size=10) if is_self else Font(name="Arial", size=10)
        c.fill      = row_fill
        c.alignment = LEFT
        apply_border(c)

        # Reg No.
        c = ws.cell(row=row, column=2)
        c.value     = m_reg
        c.font      = Font(name="Arial", italic=True, color="808080", size=10) if is_self else Font(name="Arial", size=10)
        c.fill      = row_fill
        c.alignment = CENTER
        apply_border(c)

        # Score columns C-G
        for col_i in range(3, 8):
            c = ws.cell(row=row, column=col_i)
            if is_self:
                c.value     = "N/A - Self"
                c.font      = Font(name="Arial", italic=True, color="808080", size=10)
                c.fill      = GREY_FILL
            else:
                c.value     = None
                c.font      = Font(name="Arial", size=10)
                c.fill      = row_fill
                dv.add(c)
            c.alignment = CENTER
            apply_border(c)

        # Total column H
        c = ws.cell(row=row, column=8)
        if is_self:
            c.value = "N/A"
            c.font  = Font(name="Arial", italic=True, color="808080", size=10)
            c.fill  = GREY_FILL
        else:
            c.value = "=SUM(C" + str(row) + ":G" + str(row) + ")"
            c.font  = Font(name="Arial", bold=True, size=10)
            c.fill  = GREEN_FILL
        c.alignment = CENTER
        apply_border(c)

        # Comments column I
        c = ws.cell(row=row, column=9)
        c.value     = ""
        c.font      = Font(name="Arial", italic=True, color="808080", size=10) if is_self else Font(name="Arial", size=10)
        c.fill      = GREY_FILL if is_self else row_fill
        c.alignment = LEFT
        apply_border(c)

    # Row 11 - spacer
    ws.row_dimensions[11].height = 10

    # Row 12 - Overall comments
    ws.merge_cells("A12:B12")
    c = ws["A12"]
    c.value     = "Overall Team Comments:"
    c.font      = Font(name="Arial", bold=True, size=10)
    c.alignment = LEFT

    ws.merge_cells("C12:I12")
    c = ws["C12"]
    c.value     = ""
    c.fill      = SIG_FILL
    c.alignment = LEFT
    apply_border(c)
    ws.row_dimensions[12].height = 40

    # Row 13 - spacer
    ws.row_dimensions[13].height = 8

    # Row 14 - Signature
    ws.merge_cells("A14:D14")
    c = ws["A14"]
    c.value     = "Evaluator Signature:  ___________________________"
    c.font      = Font(name="Arial", size=10)
    c.alignment = LEFT

    # Row 15 - Confidential
    ws.merge_cells("A15:I15")
    c = ws["A15"]
    c.value     = "Confidential \u2014 For Instructor Use Only"
    c.font      = Font(name="Arial", bold=True, italic=True, color="FF0000", size=10)
    c.alignment = CENTER

    # Freeze top 6 rows
    ws.freeze_panes = "A7"

wb.save("C:/ssd/Peer_Eval_SecureVault.xlsx")
print("File saved successfully.")
